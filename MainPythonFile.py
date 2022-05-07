import GeneRead
import HLS_Algorithm
import HLS_Flowchart
import HLS_Text
import PuLP_Algorithm
import ATS_Algorithm
import ATS_Flowchart
import ATS_Text
import os
import openpyxl


# Creating a specific function for finding the Minimum, Maximum and Average of an Array of tuples for each positional Value of the Tuple
# In this case the individual Tuple (,,) represents (Objective Function Value,Number of Vehicle Used,Time taken by the Heuristic)
# The function shall return Minimum, Mean, Maximum which are 3 Arays [,,] corresponsing to the same representation (Objective Function Value,Number of Vehicle Used,Time taken by the Heuristic)
def MMM(input_array_of_Tuples=[]): # This is a general function where the Tuple sizes should be limited to 3
    big_number=7777777
    minimum=[big_number,big_number,big_number]
    mean=[0,0,0]
    maximum=[-big_number,-big_number,-big_number]
    for tup in input_array_of_Tuples:
        for i in range(3): # Length of Each Tuple =3
            if tup[i]<minimum[i]:
                minimum[i]=tup[i]
            if tup[i]>maximum[i]:
                maximum[i]=tup[i]
            mean[i]=mean[i]+tup[i]
    num=len(input_array_of_Tuples)
    arithmetic_mean = [i/num for i in mean]
    return minimum,arithmetic_mean,maximum
            

directory_name="Solutions"
main_dir = directory_name
os.mkdir(main_dir)


# Call a Workbook() function of openpyxl to create a new blank Workbook object to store all the solutions for comparing with Table 3
wb = openpyxl.Workbook()
# Get workbook active sheet from the active attribute
sheet = wb.active
row_number_on_main_Excel_Table=1
column_number_on_main_Excel_Table=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "No. of Nodes considered"

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Vehicle Types Allowed"

column_number_on_main_Excel_Table+=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Objective Value as Obtained from the HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Objective Value as Obtained from the HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Objective Value as Obtained from the HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Number of Vehicles Used by HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Number of Vehicles Used by HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Number of Vehicles Used by HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum CPU Time taken by the HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average CPU Time taken by the HLS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum CPU Time taken by the HLS_Algorithm"

column_number_on_main_Excel_Table+=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Objective Value as Obtained from the HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Objective Value as Obtained from the HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Objective Value as Obtained from the HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Number of Vehicles Used by HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Number of Vehicles Used by HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Number of Vehicles Used by HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum CPU Time taken by the HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average CPU Time taken by the HLS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum CPU Time taken by the HLS_Flowchart"

column_number_on_main_Excel_Table+=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Objective Value as Obtained from the HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Objective Value as Obtained from the HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Objective Value as Obtained from the HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Number of Vehicles Used by HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Number of Vehicles Used by HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Number of Vehicles Used by HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum CPU Time taken by the HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average CPU Time taken by the HLS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum CPU Time taken by the HLS_Text"

column_number_on_main_Excel_Table+=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Objective Value as Obtained from the ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Objective Value as Obtained from the ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Objective Value as Obtained from the ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Number of Vehicles Used by ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Number of Vehicles Used by ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Number of Vehicles Used by ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum CPU Time taken by the ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average CPU Time taken by the ATS_Algorithm"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum CPU Time taken by the ATS_Algorithm"

column_number_on_main_Excel_Table+=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Objective Value as Obtained from the ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Objective Value as Obtained from the ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Objective Value as Obtained from the ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Number of Vehicles Used by ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Number of Vehicles Used by ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Number of Vehicles Used by ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum CPU Time taken by the ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average CPU Time taken by the ATS_Flowchart"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum CPU Time taken by the ATS_Flowchart"

column_number_on_main_Excel_Table+=1

cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Objective Value as Obtained from the ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Objective Value as Obtained from the ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Objective Value as Obtained from the ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum Number of Vehicles Used by ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average Number of Vehicles Used by ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum Number of Vehicles Used by ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Minimum CPU Time taken by the ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Average CPU Time taken by the ATS_Text"
cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
column_number_on_main_Excel_Table+=1
cell.value = "Maximum CPU Time taken by the ATS_Text"

#cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table+1)
#cell.value = "Objective value of solution obtained from PuLP in this same time"
# To compare with PuLP we need to relax the Constraint limiting the Number of Vehicles for each Vehicle Type
wb.save("Table.xlsx")



# Generating the Combinations as per the data available in the paper 
array_of_num_of_Nodes=[7,10,15,15,20,20,35,50,50,75,75,100,150]
corresponding_Vehicle_Types_used=[(1,2,3,4),(2,),(3,),(4,),(3,),(4,),(3,),(3,),(2,),(3,),(2,),(2,),(1,2,3,4)]


instance=len(array_of_num_of_Nodes)
if instance!=len(corresponding_Vehicle_Types_used):
    for screen_space in range(33):
        print("ReCheck Instances and Vehicle Types being used for each since the array Lengths are not the Same \n")

for h in range(instance):
    main_dir=directory_name+"/"+str(array_of_num_of_Nodes[h])+" customers served with Vehicle Type "+str(corresponding_Vehicle_Types_used[h])
    os.mkdir(main_dir)

    # Creating Arrays for calculating the Minimum, Average, Maximum values of the Objective Function, the Number of Vehicles being served with and the CPU Time for each Heuristic
    Array_HLS_Algorithm=[]
    Array_HLS_Flowchart=[]
    Array_HLS_Text=[]
    Array_ATS_Algorithm=[]
    Array_ATS_Flowchart=[]
    Array_ATS_Text=[]

    # For each conbination of Number of Nodes and Vehicle Type used, we create 2 sets of instances
    for hh in range(2):
        # Creating an instance for the Number of Nodes and Vehicle Type combination
        sub_dir=main_dir+"/Instance "+str(hh+1)
        os.mkdir(sub_dir)
        sub_dir=sub_dir+"/"

        GeneRead.Generator.Node_Generator(upto_Node_Number=array_of_num_of_Nodes[h],latitude_uniform_distribution_upper_bound=50,longitude_uniform_distribution_upper_bound=50,pickup_quantity_uniform_distribution_upper_bound=100,delivery_quantity_uniform_distribution_upper_bound=100,directory_location_to_be_saved=sub_dir)
        GeneRead.Generator.Vehicle_Type_Generator(array_of_Vehicle_Type_Indexes_considered=corresponding_Vehicle_Types_used[h],directory_to_save_Vehicle_Types_considered=sub_dir)
        #GeneRead.Generator.Lp_Norm_Random_Matrix_for_each_Vehicle_Type_Generator(p=2,directory_containing_Vehicle_Types=sub_dir,directory_containing_Node_Locations_File=sub_dir,destination_to_save_Distance_Matrix=sub_dir)
        GeneRead.Generator.Lp_Norm_Random_Matrix_for_each_Vehicle_Type_Generator(p=2,Vehicle_Types=corresponding_Vehicle_Types_used[h],directory_containing_Node_Locations_File=sub_dir,destination_to_save_Distance_Matrix=sub_dir)

        # For Each of these Instances, we shall be solving 13 Times
        for hhh in range(13):
            sub_sub_dir=sub_dir+"Solution "+str(hhh+1)
            os.mkdir(sub_sub_dir)
            sub_sub_dir=sub_sub_dir+"/"

            HLS_Algorithm_Solution,HLS_Algorithm_Time=HLS_Algorithm.HLS_Algorithm(directory_details_for_saving=sub_sub_dir,directory_containing_Distance_Matrix_file=sub_dir,directory_containing_Node_Locations_file=sub_dir,directory_containing_Vehicle_Types_file=sub_dir)
            HLS_Algorithm_Objective=HLS_Algorithm_Solution[0]
            HLS_Algorithm_Vehicles=0
            for vehicle_type in HLS_Algorithm_Solution[3]:
                HLS_Algorithm_Vehicles+=HLS_Algorithm_Solution[3][vehicle_type]
            Array_HLS_Algorithm.append((HLS_Algorithm_Objective,HLS_Algorithm_Vehicles,HLS_Algorithm_Time))

            HLS_Flowchart_Solution,HLS_Flowchart_Time=HLS_Flowchart.HLS_Flowchart(directory_details_for_saving=sub_sub_dir,directory_containing_Distance_Matrix_file=sub_dir,directory_containing_Node_Locations_file=sub_dir,directory_containing_Vehicle_Types_file=sub_dir)
            HLS_Flowchart_Objective=HLS_Flowchart_Solution[0]
            HLS_Flowchart_Vehicles=0
            for vehicle_type in HLS_Flowchart_Solution[3]:
                HLS_Flowchart_Vehicles+=HLS_Flowchart_Solution[3][vehicle_type]
            Array_HLS_Flowchart.append((HLS_Flowchart_Objective,HLS_Flowchart_Vehicles,HLS_Flowchart_Time))

            HLS_Text_Solution,HLS_Text_Time=HLS_Text.HLS_Text(directory_details_for_saving=sub_sub_dir,directory_containing_Distance_Matrix_file=sub_dir,directory_containing_Node_Locations_file=sub_dir,directory_containing_Vehicle_Types_file=sub_dir)
            HLS_Text_Objective=HLS_Text_Solution[0]
            HLS_Text_Vehicles=0
            for vehicle_type in HLS_Text_Solution[3]:
                HLS_Text_Vehicles+=HLS_Text_Solution[3][vehicle_type]
            Array_HLS_Text.append((HLS_Text_Objective,HLS_Text_Vehicles,HLS_Text_Time))

            ATS_Algorithm_Solution,ATS_Algorithm_Time=ATS_Algorithm.ATS_Algorithm(directory_details_for_saving=sub_sub_dir,directory_containing_Distance_Matrix_file=sub_dir,directory_containing_Node_Locations_file=sub_dir,directory_containing_Vehicle_Types_file=sub_dir)
            ATS_Algorithm_Objective=ATS_Algorithm_Solution[0]
            ATS_Algorithm_Vehicles=0
            for vehicle_type in ATS_Algorithm_Solution[3]:
                ATS_Algorithm_Vehicles+=ATS_Algorithm_Solution[3][vehicle_type]
            Array_ATS_Algorithm.append((ATS_Algorithm_Objective,ATS_Algorithm_Vehicles,ATS_Algorithm_Time))

            ATS_Flowchart_Solution,ATS_Flowchart_Time=ATS_Flowchart.ATS_Flowchart(directory_details_for_saving=sub_sub_dir,directory_containing_Distance_Matrix_file=sub_dir,directory_containing_Node_Locations_file=sub_dir,directory_containing_Vehicle_Types_file=sub_dir)
            ATS_Flowchart_Objective=ATS_Flowchart_Solution[0]
            ATS_Flowchart_Vehicles=0
            for vehicle_type in ATS_Flowchart_Solution[3]:
                ATS_Flowchart_Vehicles+=ATS_Flowchart_Solution[3][vehicle_type]
            Array_ATS_Flowchart.append((ATS_Flowchart_Objective,ATS_Flowchart_Vehicles,ATS_Flowchart_Time))

            ATS_Text_Solution,ATS_Text_Time=ATS_Text.ATS_Text(directory_details_for_saving=sub_sub_dir,directory_containing_Distance_Matrix_file=sub_dir,directory_containing_Node_Locations_file=sub_dir,directory_containing_Vehicle_Types_file=sub_dir)
            ATS_Text_Objective=ATS_Text_Solution[0]
            ATS_Text_Vehicles=0
            for vehicle_type in ATS_Text_Solution[3]:
                ATS_Text_Vehicles+=ATS_Text_Solution[3][vehicle_type]
            Array_ATS_Text.append((ATS_Text_Objective,ATS_Text_Vehicles,ATS_Text_Time))


    # Writing in the Original Excel Table
    row_number_on_main_Excel_Table+=1
    column_number_on_main_Excel_Table=1

    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = str(array_of_num_of_Nodes[h])

    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = str(corresponding_Vehicle_Types_used[h])


    # Finding the Minimum, Maximum and Mean of the Arrays for the Vehicle and Node Combinations
    HLS_Algorithm_Minimum,HLS_Algorithm_Mean,HLS_Algorithm_Maximum=MMM(Array_HLS_Algorithm)
    column_number_on_main_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Minimum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Mean[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Maximum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Minimum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Mean[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Maximum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Minimum[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Mean[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Algorithm_Maximum[2]

    HLS_Flowchart_Minimum,HLS_Flowchart_Mean,HLS_Flowchart_Maximum=MMM(Array_HLS_Flowchart)
    column_number_on_main_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Minimum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Mean[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Maximum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Minimum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Mean[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Maximum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Minimum[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Mean[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Flowchart_Maximum[2]

    HLS_Text_Minimum,HLS_Text_Mean,HLS_Text_Maximum=MMM(Array_HLS_Text)
    column_number_on_main_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Minimum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Mean[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Maximum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Minimum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Mean[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Maximum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Minimum[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Mean[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = HLS_Text_Maximum[2]


    ATS_Algorithm_Minimum,ATS_Algorithm_Mean,ATS_Algorithm_Maximum=MMM(Array_ATS_Algorithm)
    column_number_on_main_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Minimum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Mean[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Maximum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Minimum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Mean[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Maximum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Minimum[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Mean[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Algorithm_Maximum[2]

    ATS_Flowchart_Minimum,ATS_Flowchart_Mean,ATS_Flowchart_Maximum=MMM(Array_ATS_Flowchart)
    column_number_on_main_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Minimum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Mean[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Maximum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Minimum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Mean[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Maximum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Minimum[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Mean[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Flowchart_Maximum[2]

    ATS_Text_Minimum,ATS_Text_Mean,ATS_Text_Maximum=MMM(Array_ATS_Text)
    column_number_on_main_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Minimum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Mean[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Maximum[0]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Minimum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Mean[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Maximum[1]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Minimum[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Mean[2]
    cell = sheet.cell(row = row_number_on_main_Excel_Table, column = column_number_on_main_Excel_Table)
    column_number_on_main_Excel_Table+=1
    cell.value = ATS_Text_Maximum[2]

    wb.save("Table.xlsx")


# Cannot Compare with Heuristic and PuLP since PuLP uses VN but Heuristic has unlimited vehicles for each Type
#objec_val,Vehicle_Type_Maximum_Utilised_Capacity,Number_of_Vehicle_used_of_each_Type=PuLP_Algorithm.PuLP_Algorithm(max_seconds_allowed_for_calculation=delta_T)

for row in sheet.iter_rows():
    for cell in row:      
        cell.alignment =  cell.alignment.copy(wrapText=True)
wb.save("Table.xlsx")